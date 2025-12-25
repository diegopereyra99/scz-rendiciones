from __future__ import annotations

import io
import os
from datetime import datetime, timedelta, timezone
from typing import List

from google.auth import default as google_auth_default  # type: ignore
from google.auth.transport.requests import Request as GoogleAuthRequest  # type: ignore
from googleapiclient.discovery import build  # type: ignore
from googleapiclient.http import MediaIoBaseUpload  # type: ignore
from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter
from PIL import Image

from .. import gcs
from ..config import Settings
from ..fetch import fetch_bytes
from ..models import (
    ErrorPayload,
    FinalizeArtifact,
    FinalizeRequest,
    FinalizeResponse,
    Warning,
)
from ..utils import SUPPORTED_IMAGE_EXTS, ensure_rgb


def _drive_service():
    creds, _ = google_auth_default(scopes=["https://www.googleapis.com/auth/drive.file"])
    if creds and creds.expired and creds.refresh_token:
        creds.refresh(GoogleAuthRequest())
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def _upload_drive_file(name: str, folder_id: str, data: bytes, mime_type: str) -> str:
    service = _drive_service()
    media = MediaIoBaseUpload(io.BytesIO(data), mimetype=mime_type, resumable=False)
    metadata = {"name": name, "parents": [folder_id]}
    file = service.files().create(body=metadata, media_body=media, fields="id").execute()
    return file["id"]


def _image_bytes_to_pdf_page(image_bytes: bytes) -> bytes:
    image = Image.open(io.BytesIO(image_bytes))
    image = ensure_rgb(image)
    buf = io.BytesIO()
    image.save(buf, format="PDF")
    return buf.getvalue()


def _load_template_bytes(template_ref, settings: Settings) -> bytes:
    """
    If template_ref is None, load from embedded path (settings.xlsm_template_path).
    Otherwise fetch via GCS/Drive/signed URL.
    """
    if template_ref is None:
        path = settings.xlsm_template_path
        if not os.path.exists(path):
            raise FileNotFoundError(f"Embedded template not found at {path}")
        with open(path, "rb") as f:
            return f.read()
    return fetch_bytes(template_ref, settings)


async def run_finalize(request: FinalizeRequest, settings: Settings) -> FinalizeResponse:
    warnings: List[Warning] = []

    pdf_name = request.options.pdfName if request.options else "rendicion.pdf"
    xlsm_name = request.options.xlsmName if request.options else "rendicion.xlsm"
    signed_ttl = request.options.signedUrlTtlSeconds if request.options else settings.default_signed_url_ttl

    # Fetch cover
    try:
        cover_bytes = fetch_bytes(request.inputs.cover.model_dump(), settings)
    except Exception as exc:  # noqa: BLE001
        return FinalizeResponse(
            ok=False,
            rendicionId=request.rendicionId,
            pdf=FinalizeArtifact(),
            xlsm=FinalizeArtifact(),
            warnings=[],
            error=ErrorPayload(code="INVALID_ARGUMENT", message=f"Failed to fetch cover: {exc}", details={}),
        )

    # Build PDF
    writer = PdfWriter()

    def append_pdf_content(pdf_bytes: bytes):
        reader = PdfReader(io.BytesIO(pdf_bytes))
        for page in reader.pages:
            writer.add_page(page)

    try:
        append_pdf_content(cover_bytes)
    except Exception as exc:  # noqa: BLE001
        return FinalizeResponse(
            ok=False,
            rendicionId=request.rendicionId,
            pdf=FinalizeArtifact(),
            xlsm=FinalizeArtifact(),
            warnings=[],
            error=ErrorPayload(code="PDF_MERGE_FAILED", message=f"Cover merge failed: {exc}", details={}),
        )

    for item in request.inputs.normalizedItems:
        try:
            content = fetch_bytes(item.model_dump(), settings)
            ext = ""
            if item.mime:
                ext = item.mime.split("/")[-1]
            elif "." in item.gcsUri:
                ext = item.gcsUri.split(".")[-1].lower()
            if item.mime == "application/pdf" or ext == "pdf":
                append_pdf_content(content)
            elif (item.mime and item.mime.startswith("image/")) or ext in SUPPORTED_IMAGE_EXTS:
                pdf_page = _image_bytes_to_pdf_page(content)
                append_pdf_content(pdf_page)
            else:
                warnings.append(
                    Warning(
                        code="UNSUPPORTED_FILE_TYPE",
                        message=f"Skipping unsupported item {item.gcsUri}",
                        details={"mime": item.mime},
                    )
                )
        except Exception as exc:  # noqa: BLE001
            warnings.append(
                Warning(
                    code="PDF_MERGE_FAILED",
                    message=f"Failed to merge item {item.gcsUri}",
                    details={"error": str(exc)},
                )
            )

    pdf_buf = io.BytesIO()
    writer.write(pdf_buf)
    final_pdf_bytes = pdf_buf.getvalue()

    # Build XLSM
    try:
        template_bytes = _load_template_bytes(
            request.inputs.xlsmTemplate.model_dump() if request.inputs.xlsmTemplate else None, settings
        )
        wb = load_workbook(io.BytesIO(template_bytes), keep_vba=True)
        for cell_value in request.inputs.xlsmValues:
            ws = wb[cell_value.sheet]
            ws.cell(row=cell_value.row, column=cell_value.col).value = cell_value.value
        xlsm_buf = io.BytesIO()
        wb.save(xlsm_buf)
        final_xlsm_bytes = xlsm_buf.getvalue()
    except Exception as exc:  # noqa: BLE001
        return FinalizeResponse(
            ok=False,
            rendicionId=request.rendicionId,
            pdf=FinalizeArtifact(),
            xlsm=FinalizeArtifact(),
            warnings=warnings or None,
            error=ErrorPayload(code="XLSM_WRITE_FAILED", message=str(exc), details={}),
        )

    pdf_artifact = FinalizeArtifact()
    xlsm_artifact = FinalizeArtifact()

    try:
        if request.output.gcsPrefix:
            prefix = gcs.normalize_prefix(request.output.gcsPrefix)
            pdf_uri = gcs.upload_bytes(
                final_pdf_bytes, f"{prefix}outputs/{pdf_name}", content_type="application/pdf"
            )
            xlsm_uri = gcs.upload_bytes(
                final_xlsm_bytes,
                f"{prefix}outputs/{xlsm_name}",
                content_type="application/vnd.ms-excel.sheet.macroEnabled.12",
            )
            pdf_artifact.gcsUri = pdf_uri
            xlsm_artifact.gcsUri = xlsm_uri
            try:
                pdf_artifact.signedUrl = gcs.maybe_signed_url(pdf_uri, signed_ttl)
            except Exception as exc:  # noqa: BLE001
                warnings.append(
                    Warning(
                        code="SIGNED_URL_FAILED",
                        message="Failed to generate signed URL for PDF",
                        details={"error": str(exc)},
                    )
                )
            try:
                xlsm_artifact.signedUrl = gcs.maybe_signed_url(xlsm_uri, signed_ttl)
            except Exception as exc:  # noqa: BLE001
                warnings.append(
                    Warning(
                        code="SIGNED_URL_FAILED",
                        message="Failed to generate signed URL for XLSM",
                        details={"error": str(exc)},
                    )
                )
        elif request.output.driveFolderId:
            if not settings.drive_enabled:
                return FinalizeResponse(
                    ok=False,
                    rendicionId=request.rendicionId,
                    pdf=FinalizeArtifact(),
                    xlsm=FinalizeArtifact(),
                    warnings=warnings or None,
                    error=ErrorPayload(
                        code="DRIVE_API_DISABLED",
                        message="Drive API is disabled; enable REN_DRIVE_ENABLED to true and configure ADC.",
                        details={},
                    ),
                )
            pdf_drive_id = _upload_drive_file(pdf_name, request.output.driveFolderId, final_pdf_bytes, "application/pdf")
            xlsm_drive_id = _upload_drive_file(
                xlsm_name,
                request.output.driveFolderId,
                final_xlsm_bytes,
                "application/vnd.ms-excel.sheet.macroEnabled.12",
            )
            pdf_artifact.driveFileId = pdf_drive_id
            xlsm_artifact.driveFileId = xlsm_drive_id
    except Exception as exc:  # noqa: BLE001
        return FinalizeResponse(
            ok=False,
            rendicionId=request.rendicionId,
            pdf=FinalizeArtifact(),
            xlsm=FinalizeArtifact(),
            warnings=warnings or None,
            error=ErrorPayload(code="GCS_WRITE_FAILED", message=str(exc), details={}),
        )

    return FinalizeResponse(
        ok=True,
        rendicionId=request.rendicionId,
        pdf=pdf_artifact,
        xlsm=xlsm_artifact,
        warnings=warnings or None,
        error=None,
    )
